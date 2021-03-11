# How to modify excel files

from openpyxl import load_workbook

# workbook = Workbook()
wb = load_workbook(
    filename="D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Report table.xlsx")
sheet = wb.active

sheet['A1'] = "Hello"
sheet['B1'] = "World"

wb.save(filename="Report table.xlsx")
