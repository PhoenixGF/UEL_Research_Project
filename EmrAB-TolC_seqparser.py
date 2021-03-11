""" 

# Creates folders for each species that has the protein and writes the fasta file in the folder
import os

with open(
        r'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/TolC_1000Seq_Compressed.fasta') as f:
    fdict = {}
    fkey = ''
    for line in f:
        line = line.strip()
        if '>' in line:
            fkey = line
            fdict[fkey] = ''
        else:
            fdict[fkey] += line

proteincount = 0
for key, val in fdict.items():
    proteincount += 1

    pos1 = key.find('OS=')
    pos2 = key.find('OX=')

    dirname = f'{key[pos1+3:pos2-1]} EmrB'
    dirname = dirname.replace('/', '')
    dirname = dirname.replace(':', '-')

    filename = f'TolC {key[pos1+3:pos2-1]}, {proteincount}.fasta'
    filename = filename.replace('/', '')
    filename = filename.replace(':', '-')

    proteinpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrB'
    if not os.path.exists(proteinpath):
        os.makedirs(proteinpath)
    else:
        pass

    dirpath = f'{proteinpath}/{dirname}'
    if not os.path.exists(dirpath):
        os.makedirs(dirpath)
    else:
        pass

    filepath = f'{proteinpath}/{dirname}/{filename}'
    with open(filepath, 'w') as outputfile:
        outputfile.write(f'{key}\n{val[:60]}\n{val[60:120]}\n{val[120:180]}\n{val[180:240]}\n'
                         f'{val[240:300]}\n{val[300:360]}\n{val[360:420]}\n{val[420:480]}\n{val[480:]}\n')

print(proteincount)

# Creates a folder for each genus of the bacteria to add the species in.
import os
proteinpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrB'
output = [folder for folder in os.listdir(proteinpath) if os.path.isdir(os.path.join(proteinpath, folder))]
print(output)
for words in output:
    word = words.split()
    genus = word[0]
    newpath = f'{proteinpath}/{genus}'
    print(word, genus, newpath)
    if not os.path.exists(newpath):
        os.makedirs(newpath)
    else:
        pass

"""

# if not os.path.exists(speciespath):
#    os.makedirs(speciespath)
# else:
#    pass

# ------------------------------------------------------------------------------

"""
# Writes TolC fasta files in folders directly into the existing bacterial species folders 
import os

with open(
        r'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/TolC_1000Seq_Compressed.fasta') as f:
    fdict = {}
    fkey = ''
    for line in f:
        line = line.strip()
        if '>' in line:
            fkey = line
            fdict[fkey] = ''
        else:
            fdict[fkey] += line

speciespath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
output = [folder for folder in os.listdir(speciespath) if os.path.isdir(os.path.join(speciespath, folder))]

proteincount = 0
for key, val in fdict.items():
    proteincount += 1

    pos1 = key.find('OS=')
    pos2 = key.find('OX=')

    filename = f'TolC {key[pos1 + 3:pos2 - 1]}, {proteincount}.fasta'
    filename = filename.replace('/', '')
    filename = filename.replace(':', '-')

    foldername = f'{key[pos1 + 3:pos2 - 1]} TolC'
    foldername = foldername.replace('/', '')
    foldername = foldername.replace(':', '-')

    genus = foldername.split()
    genusfoldername = genus[0]
    print(genusfoldername)

    genusfolderpath = f'{speciespath}/{genusfoldername}'
    if not os.path.exists(genusfolderpath):
        os.makedirs(genusfolderpath)

    spfolderpath = f'{speciespath}/{genusfoldername}/{foldername}'
    if not os.path.exists(spfolderpath):
        os.makedirs(spfolderpath)

    filepath = f'{speciespath}/{genusfoldername}/{foldername}/{filename}'
    with open(filepath, 'w') as outputfile:
        outputfile.write(f'{key}\n{val[:60]}\n{val[60:120]}\n{val[120:180]}\n{val[180:240]}\n'
                         f'{val[240:300]}\n{val[300:360]}\n{val[360:420]}\n{val[420:480]}\n{val[480:]}\n')

print(proteincount)
"""

# ------------------------------------------------------------------------------

"""
import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)

wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Report table.xlsx")
sheet = wb.active

filename = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Report File Compressed.txt'
filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

rownum = 44
with open(filename, 'w') as outputfile:
    for file in os.listdir(filefolderpath):
        filepath = f'{filefolderpath}/{file}'
        with open(filepath) as f:
            fdict = {}
            fkey = ''
            for line in f:
                line = line.strip()
                if '>' in line:
                    fkey = line
                    fdict[fkey] = ''
                else:
                    fdict[fkey] += line

            seqlength = len(fdict[fkey])

        if 'EmrA' in file:
            pos1 = file.find('EmrA')
            species = file[:pos1]
            protein = 'EmrA'
            procolumn = 3
            seqcolumn = 6
        elif 'EmrB' in file:
            pos1 = file.find('EmrB')
            species = file[:pos1]
            protein = 'EmrB'
            procolumn = 2
            seqcolumn = 5
        else:
            pos1 = file.find('TolC')
            species = file[:pos1]
            protein = 'skip'
            procolumn = 4
            seqcolumn = 7

        pos2 = file.find('.fasta')

        ranking = int(file[pos1 + 4:pos2])
        xlranking = sheet.cell(row=rownum, column=11).value

        if species is not None:
            species = species.strip()
            words = species.split()
            genus = words[0]

        xlspecies = sheet.cell(row=rownum, column=1).value


        saveranking = True
        savefile = True
        if xlspecies is not None:
            xlspecies = xlspecies.strip()
            morewords = xlspecies.split()
            xlgenus = morewords[0]
            if species == xlspecies:
                if 'TolC' in file:
                    pos1 = file.find('TolC')
                    species = file[:pos1]
                    protein = 'TolC?'
                    procolumn = 4
                    seqcolumn = 7
                if xlranking is not None:
                    if int(xlranking) < ranking:
                        saveranking = False
                        savefile = False
            else:
                rownum += 1

            if genus == xlgenus:
                pass
            else:
                rownum += 1

        if saveranking is True:
            sheet.cell(row=rownum, column=11).value = ranking
        if protein != 'skip':
            sheet.cell(row=rownum, column=1).value = species
            sheet.cell(row=rownum, column=procolumn).value = protein
            sheet.cell(row=rownum, column=seqcolumn).value = seqlength
            sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'
            outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

            # outputfile.write(f'{fkey}\n{fdict[fkey][:60]}\n{fdict[fkey][60:120]}\n{fdict[fkey][120:180]}\n'
            #                  f'{fdict[fkey][180:240]}\n{fdict[fkey][240:300]}\n{fdict[fkey][300:360]}\n'
            #                  f'{fdict[fkey][360:420]}\n{fdict[fkey][420:480]}\n{fdict[fkey][480:]}\n')

    wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/New Report Table.xlsx')
"""

# save = True
#        if xlspecies is not None:
#            xlspecies = xlspecies.strip()
#            morewords = xlspecies.split()
#            xlgenus = morewords[0]
#            if species == xlspecies:
#                if 'TolC' in file:
#                    pos1 = file.find('TolC')
#                    species = file[:pos1]
#                    protein = 'TolC?'
#                    procolumn = 4
#                    seqcolumn = 7
#                if xlranking is not None:
#                    if int(xlranking) < ranking:
#                        save = False
#            else:
#                rownum += 1
#
#            if genus == xlgenus:
#                pass
#            else:
#                rownum += 1
#
#        if protein != 'skip' and save is True:
#            sheet.cell(row=rownum, column=1).value = species
#            sheet.cell(row=rownum, column=procolumn).value = protein
#            sheet.cell(row=rownum, column=seqcolumn).value = seqlength
#            sheet.cell(row=rownum, column=11).value = ranking
#            sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'


# species = species.strip()
# words = species.split()
# genus = words[0]
#
# xlspecies = sheet.cell(row=rownum - 1, column=1).value
# if xlspecies != None:
#     xlspecies = xlspecies.strip()
#     if species == xlspecies:
#         continue
#     morewords = xlspecies.split()
#     xlgenus = morewords[0]
#     if genus != xlgenus:
#         rownum += 1
# else:
#     rownum += 1

# -----------------------------------------------------------------------------------------------
"""
import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)

wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Report table.xlsx")
sheet = wb.active

filename = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Report File Compressed.txt'
filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

rownum = 44
with open(filename, 'w') as outputfile:
    for file in os.listdir(filefolderpath):
        filepath = f'{filefolderpath}/{file}'
        with open(filepath) as f:
            fdict = {}
            fkey = ''
            for line in f:
                line = line.strip()
                if '>' in line:
                    fkey = line
                    fdict[fkey] = ''
                else:
                    fdict[fkey] += line

            seqlength = len(fdict[fkey])

        if 'EmrA' in file:
            pos1 = file.find('EmrA')
            species = file[:pos1]
            protein = 'EmrA'
            procolumn = 3
            seqcolumn = 6
        elif 'EmrB' in file:
            pos1 = file.find('EmrB')
            species = file[:pos1]
            protein = 'EmrB'
            procolumn = 2
            seqcolumn = 5
        else:
            pos1 = file.find('TolC')
            species = file[:pos1]
            protein = 'skip'
            procolumn = 4
            seqcolumn = 7

        pos2 = file.find('.fasta')

        ranking = int(file[pos1 + 4:pos2])
        xlranking = sheet.cell(row=rownum, column=11).value

        if species is not None:
            species = species.strip()
            words = species.split()
            genus = words[0]

        xlspecies = sheet.cell(row=rownum, column=1).value


        saveranking = True
        savefile = True
        if xlspecies is not None:
            xlspecies = xlspecies.strip()
            morewords = xlspecies.split()
            xlgenus = morewords[0]
            if species == xlspecies:
                if 'TolC' in file:
                    pos1 = file.find('TolC')
                    species = file[:pos1]
                    protein = 'TolC?'
                    procolumn = 4
                    seqcolumn = 7
                if xlranking is not None:
                    if int(xlranking) < ranking:
                        saveranking = False
                        savefile = False
            else:
                rownum += 1

            if genus == xlgenus:
                pass
            else:
                rownum += 1

        if saveranking is True:
            sheet.cell(row=rownum, column=11).value = ranking
        if protein != 'skip':
            sheet.cell(row=rownum, column=1).value = species
            sheet.cell(row=rownum, column=procolumn).value = protein
            sheet.cell(row=rownum, column=seqcolumn).value = seqlength
            sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'
            outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

            # outputfile.write(f'{fkey}\n{fdict[fkey][:60]}\n{fdict[fkey][60:120]}\n{fdict[fkey][120:180]}\n'
            #                  f'{fdict[fkey][180:240]}\n{fdict[fkey][240:300]}\n{fdict[fkey][300:360]}\n'
            #                  f'{fdict[fkey][360:420]}\n{fdict[fkey][420:480]}\n{fdict[fkey][480:]}\n')

    wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/New Report Table.xlsx')


"""
# ------------------------------------------------------------------------------
"""

# Saves in a single file
# Does the job, but some of the code is redundant (counters could be replaced with lists) and could be optimised.

import os
from collections import Counter
import math

filename = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/' \
           'Sequence files/Report File Uncompressed.fasta'

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
           "Summaries/Identity files/EmrAB-TolC_identity.txt"

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

filelist = os.listdir(filefolderpath)
#print(filelist)
specieslist = []
onlyspecieslist = []
rankinglist = []
indexlist = []
indexcount = 0
emracounter = Counter()
emrbcounter = Counter()
tolccounter = Counter()
for file in filelist:
    if 'EmrA' in file:
        filepos1 = file.find('EmrA')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in emracounter:
            emracounter[onlyspecies] += 1
    elif 'EmrB' in file:
        filepos1 = file.find('EmrB')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in emrbcounter:
            emrbcounter[onlyspecies] += 1
    else:
        filepos1 = file.find('TolC')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in tolccounter:
            tolccounter[onlyspecies] += 1

    filepos2 = file.find('.fasta')

    if species not in specieslist:
        specieslist += [species]

        onlyspecieslist += [onlyspecies]

        ranking = file[filepos1 + 5:filepos2]
        rankinglist += [ranking]

        indexlist += str(indexcount)
        indexcount += 1

# print(emracounter)
# print(emrbcounter)
# print(tolccounter)
# print(onlyspecieslist)

# filters each filename and creates a dictionary of files according to criteria:
# if tolc is on its own, don't save; if ranking is lower, don't save;

sprankdict = {}
# sprankkey = ''
for species, justspecies, ranking in zip(specieslist, onlyspecieslist, rankinglist):
    if 'TolC' in species:
        #print(species)
        if justspecies not in emracounter and emrbcounter:
            # print(species)
            pass
        else:
            sprankkey = species
            sprankdict[sprankkey] = ranking
    else:
        sprankkey = species
        sprankdict[sprankkey] = ranking
# print(sprankdict)

# converts the dictionary of files with ranking into a list of file names with name and ranking together
newfilename = []
for key, val in sprankdict.items():
    newfilename += [' '.join([key, val])]
    # print(newfilename)

# compiles identity numbers from each protein into a dictionary
with open(ifilename) as outputfile:
    idict = {}
    ikey = ''
    for eachline in outputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            if eachline[0].isdigit() is True:
                idict[ikey] = eachline
            else:
                ikey = spline[0]

# reads through the list of filtered files and writes them all into a single file with some features
with open(filename, 'w') as outputfile:
    for newfile in newfilename:
        filepath = f'{filefolderpath}/{newfile}.fasta'
        with open(filepath) as f:
            fdict = {}
            fkey = ''
            for line in f:
                line = line.strip()
                if '>' in line:
                    fkey = line
                    fdict[fkey] = ''
                else:
                    fdict[fkey] += line

            seqlength = len(fdict[fkey])

            numberpos1 = fkey.find('|')
            numberpos2 = fkey.rfind('|')
            uniprotnumber = fkey[numberpos1 + 1:numberpos2]

            # gets the identity in of the protein from the dictionary, if it is there
            if uniprotnumber in idict:
                identity = idict[uniprotnumber]
            else:
                identity = 'identity not found'

        # could add ranking by removing -4
        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n{newfile[:-4]} Sequence length {seqlength}\n\n')

        timesby = math.trunc(seqlength / 60)
        start = 0
        end = 60
        outputfile.write(f'{fkey}\n')
        for i in range(timesby):
            outputfile.write(f'{fdict[fkey][start:end]}\n')
            start += 60
            end += 60
        outputfile.write(f'{fdict[fkey][start:]}\n{newfile[:-4]}'
                         f' Sequence length {seqlength} Identity {identity}\n\n')

"""


"""
# Saves in separate files
import os
from collections import Counter
import math

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

newfilefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/' \
                    'Research Project/Summaries/EmrAB-TolC Compressed'

filelist = os.listdir(filefolderpath)
print(filelist)
specieslist = []
onlyspecieslist = []
rankinglist = []
indexlist = []
indexcount = 0
emracounter = Counter()
emrbcounter = Counter()
tolccounter = Counter()
for file in filelist:
    filepos2 = file.find('.fasta')
    if 'EmrA' in file:
        filepos1 = file.find('EmrA')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in emracounter:
            emracounter[onlyspecies] += 1
    elif 'EmrB' in file:
        filepos1 = file.find('EmrB')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in emrbcounter:
            emrbcounter[onlyspecies] += 1
    else:
        filepos1 = file.find('TolC')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        if species not in tolccounter:
            tolccounter[onlyspecies] += 1

    if species not in specieslist:
        specieslist += [species]

        onlyspecieslist += [onlyspecies]

        ranking = file[filepos1 + 5:filepos2]
        rankinglist += [ranking]

        indexlist += str(indexcount)
        indexcount += 1

print(emracounter)
print(emrbcounter)
print(tolccounter)
print(specieslist)
print(onlyspecieslist)

sprankdict = {}
# sprankkey = ''
for species, justspecies, ranking in zip(specieslist, onlyspecieslist, rankinglist):
    if 'TolC' in species:
        #print(species)
        if justspecies not in emracounter and emrbcounter:
            # print(species)
            pass
        else:
            sprankkey = species
            sprankdict[sprankkey] = ranking
    else:
        sprankkey = species
        sprankdict[sprankkey] = ranking
print(sprankdict)


newfilename = []
for key, val in sprankdict.items():
    newfilename += [' '.join([key, val])]
    #print(newfilename)


for newfile in newfilename:
    newfilepath = f'{newfilefolderpath}/{newfile}.fasta'
    with open(newfilepath, 'w') as outputfile:
        filepath = f'{filefolderpath}/{newfile}.fasta'
        with open(filepath) as f:
                fdict = {}
                fkey = ''
                for line in f:
                    line = line.strip()
                    if '>' in line:
                        fkey = line
                        fdict[fkey] = ''
                    else:
                        fdict[fkey] += line

                seqlength = len(fdict[fkey])

        # could add ranking by removing -4
        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n{newfile[:-4]} Sequence length {seqlength}\n\n')

        timesby = math.trunc(seqlength / 60)
        start = 0
        end = 60
        outputfile.write(f'{fkey}\n')
        for i in range(timesby):
            outputfile.write(f'{fdict[fkey][start:end]}\n')
            start += 60
            end += 60
        outputfile.write(f'{fdict[fkey][start:]}\n')

"""
# -------------------------------------------------------------------------------------------
"""
# Writes in Excel with identities, but only EmrB or EmrA's.
import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Report table.xlsx")
sheet = wb.active

filename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
           "Summaries/Identity files/EmrAB-TolC_identity.txt"
filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

rownum = 44
with open(filename) as outputfile:
    idict = {}
    ikey = ''
    for eachline in outputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            if eachline[0].isdigit() is True:
                idict[ikey] = eachline
            else:
                ikey = spline[0]

# can add count = 0 to count how many uniprotnumber in idict or not in idict (1000 and 2000)

for file in os.listdir(filefolderpath):
    filepath = f'{filefolderpath}/{file}'
    with open(filepath) as f:
        fdict = {}
        fkey = ''
        for line in f:
            line = line.strip()
            if '>' in line:
                fkey = line
                fdict[fkey] = ''
            else:
                fdict[fkey] += line

    numberpos1 = fkey.find('|')
    numberpos2 = fkey.rfind('|')
    uniprotnumber = fkey[numberpos1 + 1:numberpos2]

    seqlength = len(fdict[fkey])

    if uniprotnumber in idict:
        identity = idict[uniprotnumber]
    else:
        identity = 'identity not found'

    if 'EmrA' in file:
        pos1 = file.find('EmrA')
        species = file[:pos1]
        protein = 'EmrA'
        procolumn = 3
        seqcolumn = 6
    elif 'EmrB' in file:
        pos1 = file.find('EmrB')
        species = file[:pos1]
        protein = 'EmrB'
        procolumn = 2
        seqcolumn = 5
    else:
        pos1 = file.find('TolC')
        species = file[:pos1]
        protein = 'skip'
        procolumn = 4
        seqcolumn = 7

    pos2 = file.find('.fasta')

    # ranking = int(file[pos1 + 4:pos2])
    # xlranking = sheet.cell(row=rownum, column=11).value

    if species is not None:
        species = species.strip()
        words = species.split()
        genus = words[0]
    else:
        genus = ''

    xlspecies = sheet.cell(row=rownum, column=1).value

    # saveranking = True
    if xlspecies is not None:
        xlspecies = xlspecies.strip()
        morewords = xlspecies.split()
        xlgenus = morewords[0]
        if species == xlspecies:
            if 'TolC' in file:
                pos1 = file.find('TolC')
                species = file[:pos1]
                protein = 'TolC?'
                procolumn = 4
                seqcolumn = 7
            # if xlranking.isdigit == True:
            #     print(xlranking)
            #     if int(xlranking) < ranking:
            #         saveranking = False
        else:
            rownum += 1

        if genus == xlgenus:
            pass
        else:
            rownum += 1

    # if saveranking is True:
    #     sheet.cell(row=rownum, column=11).value = identity

    if protein != 'skip':
        sheet.cell(row=rownum, column=1).value = species
        sheet.cell(row=rownum, column=procolumn).value = protein
        if protein == 'EmrB':
            sheet.cell(row=rownum, column=11).value = identity
        elif protein == 'EmrA' and sheet.cell(row=rownum, column=11).value is None:
            sheet.cell(row=rownum, column=11).value = identity
        sheet.cell(row=rownum, column=seqcolumn).value = seqlength
        sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'

        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

        # outputfile.write(f'{fkey}\n{fdict[fkey][:60]}\n{fdict[fkey][60:120]}\n{fdict[fkey][120:180]}\n'
        #                  f'{fdict[fkey][180:240]}\n{fdict[fkey][240:300]}\n{fdict[fkey][300:360]}\n'
        #                  f'{fdict[fkey][360:420]}\n{fdict[fkey][420:480]}\n{fdict[fkey][480:]}\n')

wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/New Report Table.xlsx')
"""
# --------------------------------------------------------------------------------------


"""

import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Report table.xlsx")
sheet = wb.active

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
           "Summaries/Identity files/EmrAB-TolC_identity.txt"

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC Compressed'


with open(ifilename) as outputfile:
    idict = {}
    ikey = ''
    for eachline in outputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            if eachline[0].isdigit() is True:
                idict[ikey] = eachline
            else:
                ikey = spline[0]

# can add count = 0 to count how many uniprotnumber in idict or not in idict (1000 and 2000)
rownum = 44
for file in os.listdir(filefolderpath):
    filepath = f'{filefolderpath}/{file}'
    with open(filepath) as f:
        fdict = {}
        fkey = ''
        for line in f:
            line = line.strip()
            if '>' in line:
                fkey = line
                fdict[fkey] = ''
            else:
                fdict[fkey] += line

    numberpos1 = fkey.find('|')
    numberpos2 = fkey.rfind('|')
    uniprotnumber = fkey[numberpos1 + 1:numberpos2]

    seqlength = len(fdict[fkey])

    if uniprotnumber in idict:
        identity = idict[uniprotnumber]
    else:
        identity = 'identity not found'

    if 'EmrA' in file:
        pos1 = file.find('EmrA')
        species = file[:pos1]
        protein = 'EmrA'
        procolumn = 3
        seqcolumn = 6
    elif 'EmrB' in file:
        pos1 = file.find('EmrB')
        species = file[:pos1]
        protein = 'EmrB'
        procolumn = 2
        seqcolumn = 5
    else:
        pos1 = file.find('TolC')
        species = file[:pos1]
        protein = 'skip'
        procolumn = 4
        seqcolumn = 7

    pos2 = file.find('.fasta')


    if species is not None:
        species = species.strip()
        words = species.split()
        genus = words[0]
    else:
        genus = ''

    xlspecies = sheet.cell(row=rownum, column=1).value

    if xlspecies is not None:
        xlspecies = xlspecies.strip()
        morewords = xlspecies.split()
        xlgenus = morewords[0]
        if species == xlspecies:
            if 'TolC' in file:
                pos1 = file.find('TolC')
                species = file[:pos1]
                protein = 'TolC?'
                procolumn = 4
                seqcolumn = 7
        else:
            rownum += 1
        if genus == xlgenus:
            pass
        else:
            rownum += 1

    if protein != 'skip':
        sheet.cell(row=rownum, column=1).value = species
        sheet.cell(row=rownum, column=procolumn).value = protein
        if sheet.cell(row=rownum, column=11).value is None:
            # !!! creates potential problems for future automation of parsing and rewriting
            sheet.cell(row=rownum, column=11).value = identity
        else:
            sheet.cell(row=rownum, column=11).value += f', {identity}'
        sheet.cell(row=rownum, column=seqcolumn).value = seqlength
        sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'

        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

        # needs import math -> timesby = math.trunc(seqlength / 60)
        # start = 0
        # end = 60
        # outputfile.write(f'{fkey}\n')
        # for i in range(timesby):
        #     outputfile.write(f'{fdict[fkey][start:end]}\n')
        #     start += 60
        #     end += 60
        # outputfile.write(f'{fdict[fkey][start:]}\n{newfile[:-4]}'
        #                  f' Sequence length {seqlength} Identity {identity}\n\n')

wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/New Report Table.xlsx')

"""


"""

import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Report table.xlsx")
sheet = wb.active

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
           "Summaries/Identity files/EmrAB-TolC_identity.txt"

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC Compressed'


with open(ifilename) as outputfile:
    idict = {}
    ikey = ''
    for eachline in outputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            if eachline[0].isdigit() is True:
                idict[ikey] = eachline
            else:
                ikey = spline[0]

# can add count = 0 to count how many uniprotnumber in idict or not in idict (1000 and 2000)
rownum = 44
for file in os.listdir(filefolderpath):
    filepath = f'{filefolderpath}/{file}'
    with open(filepath) as f:
        fdict = {}
        fkey = ''
        for line in f:
            line = line.strip()
            if '>' in line:
                fkey = line
                fdict[fkey] = ''
            else:
                fdict[fkey] += line

    numberpos1 = fkey.find('|')
    numberpos2 = fkey.rfind('|')
    uniprotnumber = fkey[numberpos1 + 1:numberpos2]

    seqlength = len(fdict[fkey])

    if uniprotnumber in idict:
        identity = idict[uniprotnumber]
    else:
        identity = 'identity not found'

    if 'EmrA' in file:
        pos1 = file.find('EmrA')
        species = file[:pos1]
        protein = 'EmrA'
        procolumn = 3
        seqcolumn = 6
    elif 'EmrB' in file:
        pos1 = file.find('EmrB')
        species = file[:pos1]
        protein = 'EmrB'
        procolumn = 2
        seqcolumn = 5
    else:
        pos1 = file.find('TolC')
        species = file[:pos1]
        protein = 'skip'
        procolumn = 4
        seqcolumn = 7

    pos2 = file.find('.fasta')


    if species is not None:
        species = species.strip()
        words = species.split()
        genus = words[0]
    else:
        genus = ''

    xlspecies = sheet.cell(row=rownum, column=1).value

    if xlspecies is not None:
        xlspecies = xlspecies.strip()
        morewords = xlspecies.split()
        xlgenus = morewords[0]
        if species == xlspecies:
            if 'TolC' in file:
                pos1 = file.find('TolC')
                species = file[:pos1]
                protein = 'TolC?'
                procolumn = 4
                seqcolumn = 7
        else:
            rownum += 1
        if genus == xlgenus:
            pass
        else:
            rownum += 1

    # did not find a way to incorporate this yet numidentity = float(identity[:-1])

    if protein != 'skip':
        sheet.cell(row=rownum, column=1).value = species
        sheet.cell(row=rownum, column=procolumn).value = protein
        if sheet.cell(row=rownum, column=11).value is None:
            sheet.cell(row=rownum, column=11).value = identity[:-1]
        else:
            sheet.cell(row=rownum, column=11).value += f', {identity[:-1]}'
        sheet.cell(row=rownum, column=seqcolumn).value = seqlength
        sheet.cell(row=rownum, column=12).value = 'UniProt BLAST'

        # can add float() and [:-1] to identity to record the value as a number without the percent sign
        # if protein == 'EmrB':
        #     sheet.cell(row=rownum, column=11).value = identity
        # elif protein == 'EmrA' and sheet.cell(row=rownum, column=11).value is None:
        #     sheet.cell(row=rownum, column=11).value = identity

        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

        # needs import math -> timesby = math.trunc(seqlength / 60)
        # start = 0
        # end = 60
        # outputfile.write(f'{fkey}\n')
        # for i in range(timesby):
        #     outputfile.write(f'{fdict[fkey][start:end]}\n')
        #     start += 60
        #     end += 60
        # outputfile.write(f'{fdict[fkey][start:]}\n{newfile[:-4]}'
        #                  f' Sequence length {seqlength} Identity {identity}\n\n')

wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
        'Summaries/Tables and Reports/New Report Table.xlsx')

"""
"""
# Sorts out sequences and saves in separate files
import os
import math

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/EmrAB-TolC'

newfilefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/' \
                    'Research Project/Summaries/EmrAB-TolC Compressed'

filelist = os.listdir(filefolderpath)
print(filelist)
specieslist = []
onlyspecieslist = []
rankinglist = []
indexlist = []
indexcount = 0
rankingdict = {}
rankingkey = ''
emracounter = {}
emrbcounter = {}
tolccounter = {}
for file in filelist:

    filepos2 = file.find('.fasta')
    if 'EmrA' in file:
        filepos1 = file.find('EmrA')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        ranking = file[filepos1 + 5:filepos2]
        if onlyspecies in emracounter:
            if int(emracounter[onlyspecies]) > int(ranking):
                emrakey = onlyspecies
                emracounter[emrakey] = ranking
        else:
            emrakey = onlyspecies
            emracounter[emrakey] = ranking
    elif 'EmrB' in file:
        filepos1 = file.find('EmrB')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        ranking = file[filepos1 + 5:filepos2]
        if onlyspecies in emrbcounter:
            if int(emrbcounter[onlyspecies]) > int(ranking):
                #print(species, emrbcounter[onlyspecies], ranking)
                emrbkey = onlyspecies
                emrbcounter[emrbkey] = ranking
        else:
            emrbkey = onlyspecies
            emrbcounter[emrbkey] = ranking
    else:
        filepos1 = file.find('TolC')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        ranking = file[filepos1 + 5:filepos2]
        if onlyspecies in tolccounter:
            if int(tolccounter[onlyspecies]) > int(ranking):
                tolckey = onlyspecies
                tolccounter[tolckey] = ranking
        else:
            tolckey = onlyspecies
            tolccounter[tolckey] = ranking

    if species in specieslist:
        if int(rankingdict[species]) > int(ranking):
            rankingkey = species
            rankingdict[rankingkey] = ranking
            specieslist += [species]
            onlyspecieslist += [onlyspecies]
            rankinglist += [ranking]
            indexlist += str(indexcount)
            indexcount += 1
    else:
        rankingkey = species
        rankingdict[rankingkey] = ranking
        specieslist += [species]
        onlyspecieslist += [onlyspecies]
        rankinglist += [ranking]
        indexlist += str(indexcount)
        indexcount += 1


print(rankingdict)
print(emracounter)
print(emrbcounter)
print(tolccounter)
print(specieslist)
print(onlyspecieslist)

sprankdict = {}
# sprankkey = ''
for species, justspecies, ranking in zip(specieslist, onlyspecieslist, rankinglist):
    if 'TolC' in species:
        #print(species)
        if justspecies not in emracounter and emrbcounter:
            # print(species)
            pass
        else:
            sprankkey = species
            sprankdict[sprankkey] = ranking
    else:
        sprankkey = species
        sprankdict[sprankkey] = ranking
print(sprankdict)


newfilename = []
for key, val in sprankdict.items():
    newfilename += [' '.join([key, val])]
print(newfilename)


for newfile in newfilename:
    newfilepath = f'{newfilefolderpath}/{newfile}.fasta'
    with open(newfilepath, 'w') as outputfile:
        filepath = f'{filefolderpath}/{newfile}.fasta'
        with open(filepath) as f:
                fdict = {}
                fkey = ''
                for line in f:
                    line = line.strip()
                    if '>' in line:
                        fkey = line
                        fdict[fkey] = ''
                    else:
                        fdict[fkey] += line

                seqlength = len(fdict[fkey])

        # could add ranking by removing -4
        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n{newfile[:-4]} Sequence length {seqlength}\n\n')

        timesby = math.trunc(seqlength / 60)
        start = 0
        end = 60
        outputfile.write(f'{fkey}\n')
        for i in range(timesby):
            outputfile.write(f'{fdict[fkey][start:end]}\n')
            start += 60
            end += 60
        outputfile.write(f'{fdict[fkey][start:]}\n')
        
"""
# byte = b'\xff'
# head = [next(f) for i in range(5)]
# print(head)
with open('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
          'Summaries/BLAST/blast-2.11.0+/db/pycombined.fasta') as f:
    head = [next(f) for i in range(5)]
    print(head)
# with open('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
#           'Summaries/BLAST/blast-2.11.0+/db/utf8combined.fasta', encoding='utf-8-sig') as f:
#     with open(
#             'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
#             'Summaries/BLAST/blast-2.11.0+/db/pycombined.fasta', 'w') as outputfile:
#         for line in f:
#             outputfile.write(line)
