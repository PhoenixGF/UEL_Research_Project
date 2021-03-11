"""
# Writes TolC fasta files in folders directly into the existing bacterial species folders

import os
import math

with open(
        r'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/'
        r'BLAST/blast-2.11.0+/files/tolc35forparsing.fasta') as f:
    fdict = {}
    fkey = ''
    for line in f:
        line = line.strip()
        if '>' in line:
            fkey = line
            fdict[fkey] = ''
        else:
            fdict[fkey] += line

# speciespath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [folder for folder in os.listdir(speciespath) if os.path.isdir(os.path.join(speciespath, folder))]
folderpath = r'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
             r'Summaries/BLAST/blast-2.11.0+/files/TolC Sequences'

proteincount = 0
for key, val in fdict.items():
    proteincount += 1

    seqlength = len(val)

    pos1 = key.find('OS=')
    pos2 = key.find('OX=')

    filename = f'{key[pos1 + 3 : pos2 - 1]} TolC.fasta'
    filename = filename.replace('/', '')
    filename = filename.replace(':', '-')

    spenus = filename.split()
    genus = spenus[0]
    species = str(spenus[0:2])
    #print(species)

    if not os.path.exists(folderpath):
        os.makedirs(folderpath)

    if filename not in os.listdir(folderpath):
        filepath = f'{folderpath}/{filename}'
        with open(filepath, 'w') as outputfile:
            timesby = math.trunc(seqlength / 60)
            start = 0
            end = 60
            outputfile.write(f'{key}\n')
            for i in range(timesby):
                outputfile.write(f'{val[start:end]}\n')
                start += 60
                end += 60
            outputfile.write(f'{val[start:]}\n')

print(proteincount)
"""
# -------------------------------------------------------------------------------------------------
"""

# Sorts out sequences and saves in separate files
import os
import math

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
                 'Summaries/BLAST/blast-2.11.0+/files/EmrAB-TolC Sequences'

newfilefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/' \
                    'Research Project/Summaries/BLAST/blast-2.11.0+/files/EmrAB-TolC Seq Sorted'

if not os.path.exists(newfilefolderpath):
    os.makedirs(newfilefolderpath)

filelist = os.listdir(filefolderpath)
print(filelist)
specieslist = []
onlyspecieslist = []
indexlist = []
indexcount = 0
emralist = []
emrblist = []
tolclist = []

for file in filelist:
    filepos2 = file.find('.fasta')

    if 'EmrA' in file:
        filepos1 = file.find('EmrA')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        emralist += [onlyspecies]
    elif 'EmrB' in file:
        filepos1 = file.find('EmrB')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        emrblist += [onlyspecies]
    else:
        filepos1 = file.find('TolC')
        species = file[:filepos1 + 4]
        onlyspecies = file[:filepos1-1]
        tolclist += [onlyspecies]

    specieslist += [species]
    onlyspecieslist += [onlyspecies]
    indexlist += str(indexcount)
    indexcount += 1


print(emralist)
print(emrblist)
print(tolclist)
print(specieslist)
print(onlyspecieslist)

tolcsortedlist = []
for species, justspecies in zip(specieslist, onlyspecieslist):
    if 'TolC' in species:
        # print(species)
        if justspecies not in emrblist:
            if justspecies not in emralist:
                print('Not in emrb list', species)
                pass
            else:
                tolcsortedlist += [species]
        else:
            tolcsortedlist += [species]
    else:
        tolcsortedlist += [species]
print(tolcsortedlist)


for newfile in tolcsortedlist:
    newfilepath = f'{newfilefolderpath}/{newfile}.fasta'
    with open(newfilepath, 'w') as outputfile:
        filepath = f'{filefolderpath}/{newfile}.fasta'
        with open(filepath) as f:
                fdict = {}
                fkey = ''
                for line in f:
                    line = line.strip()
                    if '>' in line:
                        fkey = line
                        fdict[fkey] = ''
                    else:
                        fdict[fkey] += line

                seqlength = len(fdict[fkey])

        # could add ranking by removing -4
        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n{newfile[:-4]} Sequence length {seqlength}\n\n')

        timesby = math.trunc(seqlength / 60)
        start = 0
        end = 60
        outputfile.write(f'{fkey}\n')
        for i in range(timesby):
            outputfile.write(f'{fdict[fkey][start:end]}\n')
            start += 60
            end += 60
        outputfile.write(f'{fdict[fkey][start:]}\n')

"""
# -------------------------------------------------------------------------------------------------------
"""

# writes in excel

import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/"
    "Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/BLAST Report table.xlsx")
sheet = wb.active

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
                 'Summaries/BLAST/blast-2.11.0+/files/EmrAB-TolC Seq Sorted'

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
            "Summaries/BLAST/blast-2.11.0+/files/emrabtolcidentity_35blast.txt"

with open(ifilename) as outputfile:
    idict = {}
    ikey = ''
    for eachline in outputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            ikey = spline[0]
            idict[ikey] = spline[1]
print(idict)

# can add count = 0 to count how many uniprotnumber in idict or not in idict (1000 and 2000)
rownum = 2
for file in os.listdir(filefolderpath):
    filepath = f'{filefolderpath}/{file}'
    with open(filepath) as f:
        fdict = {}
        fkey = ''
        for line in f:
            line = line.strip()
            if '>' in line:
                fkey = line
                fdict[fkey] = ''
            else:
                fdict[fkey] += line

    numberpos1 = fkey.find('>')
    numberpos2 = fkey.find(' ')
    idnumber = fkey[numberpos1 + 1:numberpos2]

    seqlength = len(fdict[fkey])

    if idnumber in idict:
        identity = idict[idnumber]
    else:
        identity = 'identity not found'

    if 'EmrA' in file:
        pos1 = file.find('EmrA')
        species = file[:pos1]
        protein = 'EmrA'
        procolumn = 3
        seqcolumn = 6
    elif 'EmrB' in file:
        pos1 = file.find('EmrB')
        species = file[:pos1]
        protein = 'EmrB'
        procolumn = 2
        seqcolumn = 5
    else:
        pos1 = file.find('TolC')
        species = file[:pos1]
        protein = 'skip'
        procolumn = 4
        seqcolumn = 7

    pos2 = file.find('.fasta')

    if species is not None:
        species = species.strip()
        words = species.split()
        genus = words[0]
    else:
        genus = ''

    xlspecies = sheet.cell(row=rownum, column=1).value

    if xlspecies is not None:
        xlspecies = xlspecies.strip()
        morewords = xlspecies.split()
        xlgenus = morewords[0]
        if species == xlspecies:
            if 'TolC' in file:
                pos1 = file.find('TolC')
                species = file[:pos1]
                protein = 'TolC?'
                procolumn = 4
                seqcolumn = 7
        else:
            rownum += 1
        if genus == xlgenus:
            pass
        else:
            rownum += 1

    # did not find a way to incorporate this yet numidentity = float(identity[:-1])

    if protein != 'skip':
        sheet.cell(row=rownum, column=1).value = species
        sheet.cell(row=rownum, column=procolumn).value = protein
        if sheet.cell(row=rownum, column=11).value is None:
            sheet.cell(row=rownum, column=11).value = identity
        else:
            sheet.cell(row=rownum, column=11).value += f', {identity}'
        sheet.cell(row=rownum, column=seqcolumn).value = seqlength
        sheet.cell(row=rownum, column=12).value = 'NCBI BLAST on UniProt Databases'

        # can add float() and [:-1] to identity to record the value as a number without the percent sign
        # if protein == 'EmrB':
        #     sheet.cell(row=rownum, column=11).value = identity
        # elif protein == 'EmrA' and sheet.cell(row=rownum, column=11).value is None:
        #     sheet.cell(row=rownum, column=11).value = identity

        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

        # needs import math -> timesby = math.trunc(seqlength / 60)
        # start = 0
        # end = 60
        # outputfile.write(f'{fkey}\n')
        # for i in range(timesby):
        #     outputfile.write(f'{fdict[fkey][start:end]}\n')
        #     start += 60
        #     end += 60
        # outputfile.write(f'{fdict[fkey][start:]}\n{newfile[:-4]}'
        #                  f' Sequence length {seqlength} Identity {identity}\n\n')

wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
        'Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/BLAST Report table.xlsx')

"""
# --------------------------------------------------------------------------------------------------------
"""
# writes sequences with the identity numbers and other values

import os
import math

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


# wb = openpyxl.load_workbook(
#     "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/"
#     "Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/BLAST Report table.xlsx")
# sheet = wb.active

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
                 'Summaries/BLAST/blast-2.11.0+/files/EmrAB-TolC Seq Sorted'

filename = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
                 'Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/Report file Uncompressed.fasta'

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
            "Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/emrabtolcidentity_35blast.txt"

with open(ifilename) as inputfile:
    idict = {}
    ikey = ''
    for eachline in inputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            ikey = spline[0]
            idict[ikey] = spline[1]
print(idict)
print(idict['P0AEJ0'])

# can add count = 0 to count how many idnumber in idict or not in idict (1000 and 2000)
with open(filename, 'w') as outputfile:
    for file in os.listdir(filefolderpath):
        filepath = f'{filefolderpath}/{file}'
        with open(filepath) as f:
            fdict = {}
            fkey = ''
            for line in f:
                line = line.strip()
                if '>' in line:
                    fkey = line
                    fdict[fkey] = ''
                else:
                    fdict[fkey] += line

        numberpos1 = fkey.find('>')
        numberpos2 = fkey.find(' ')
        idnumber = fkey[numberpos1 + 1:numberpos2]

        seqlength = len(fdict[fkey])

        if idnumber in idict:
            identity = idict[idnumber]
        else:
            identity = 'identity not found'

        timesby = math.trunc(seqlength / 60)
        start = 0
        end = 60
        outputfile.write(f'{fkey}\n')
        for i in range(timesby):
            outputfile.write(f'{fdict[fkey][start:end]}\n')
            start += 60
            end += 60
        outputfile.write(f'{fdict[fkey][start:]}\n{file[:-6]}'
                         f' Sequence length {seqlength} Identity {identity}\n\n')
"""



# -------------------------------------------------------------------------------------------------------


import openpyxl
import os

# genusfolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/Summaries/Species'
# output = [names for names in os.listdir(genusfolderpath) if os.path.isdir(os.path.join(genusfolderpath, names))]
# print(output)


wb = openpyxl.load_workbook(
    "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/"
    "Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/BLAST Report Table V2.xlsx")
sheet = wb.active

filefolderpath = 'D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/' \
                 'Summaries/BLAST/blast-2.11.0+/files/EmrAB-TolC Seq Sorted'

ifilename = "D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/" \
            "Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/emrabtolcidentity_35blast.txt"

with open(ifilename) as inputfile:
    idict = {}
    ikey = ''
    for eachline in inputfile:
        eachline = eachline.strip()
        spline = eachline.split()
        if eachline != '':
            ikey = spline[0]
            idict[ikey] = spline[1]
print(idict)

# can add count = 0 to count how many uniprotnumber in idict or not in idict (1000 and 2000)
rownum = 2
for file in os.listdir(filefolderpath):
    filepath = f'{filefolderpath}/{file}'
    with open(filepath) as f:
        fdict = {}
        fkey = ''
        for line in f:
            line = line.strip()
            if '>' in line:
                fkey = line
                fdict[fkey] = ''
            else:
                fdict[fkey] += line

    numberpos1 = fkey.find('>')
    numberpos2 = fkey.find(' ')
    idnumber = fkey[numberpos1 + 1:numberpos2]

    seqlength = len(fdict[fkey])

    if idnumber in idict:
        identity = idict[idnumber]
    else:
        identity = 'identity not found'

    if 'EmrA' in file:
        pos1 = file.find('EmrA')
        species = file[:pos1]
        protein = 'EmrA'
        procolumn = 3
        seqcolumn = 6
        idcolumn = 9
    elif 'EmrB' in file:
        pos1 = file.find('EmrB')
        species = file[:pos1]
        protein = 'EmrB'
        procolumn = 2
        seqcolumn = 5
        idcolumn = 8
    else:
        pos1 = file.find('TolC')
        species = file[:pos1]
        protein = 'skip'
        procolumn = 4
        seqcolumn = 7
        idcolumn = 10

    pos2 = file.find('.fasta')

    if species is not None:
        species = species.strip()
        words = species.split()
        genus = words[0]
    else:
        genus = ''

    xlspecies = sheet.cell(row=rownum, column=1).value

    if xlspecies is not None:
        xlspecies = xlspecies.strip()
        morewords = xlspecies.split()
        xlgenus = morewords[0]
        if species == xlspecies:
            if 'TolC' in file:
                pos1 = file.find('TolC')
                species = file[:pos1]
                protein = 'TolC?'
                procolumn = 4
                seqcolumn = 7
                idcolumn = 10
        else:
            rownum += 1
        if genus == xlgenus:
            pass
        else:
            rownum += 1

    # did not find a way to incorporate this yet numidentity = float(identity[:-1])

    if protein != 'skip':
        sheet.cell(row=rownum, column=1).value = species
        sheet.cell(row=rownum, column=procolumn).value = protein
        sheet.cell(row=rownum, column=seqcolumn).value = seqlength
        sheet.cell(row=rownum, column=idcolumn).value = identity
        sheet.cell(row=rownum, column=15).value = 'NCBI BLAST on UniProt Databases'

        # can add float() and [:-1] to identity to record the value as a number without the percent sign
        # if protein == 'EmrB':
        #     sheet.cell(row=rownum, column=11).value = identity
        # elif protein == 'EmrA' and sheet.cell(row=rownum, column=11).value is None:
        #     sheet.cell(row=rownum, column=11).value = identity

        # outputfile.write(f'{fkey}\n{fdict[fkey]}\n\n')

        # needs import math -> timesby = math.trunc(seqlength / 60)
        # start = 0
        # end = 60
        # outputfile.write(f'{fkey}\n')
        # for i in range(timesby):
        #     outputfile.write(f'{fdict[fkey][start:end]}\n')
        #     start += 60
        #     end += 60
        # outputfile.write(f'{fdict[fkey][start:]}\n{newfile[:-4]}'
        #                  f' Sequence length {seqlength} Identity {identity}\n\n')

wb.save('D:/Phoenix/Documents Microsoft/UEL BS/Level 6/Research Project/'
        'Summaries/BLAST/blast-2.11.0+/files/Tables and Report files/BLAST Report Table V2.xlsx')


# --------------------------------------------------------------------------------------------------------